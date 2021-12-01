import pandas
import os
from flask import session
import statsmodels.tsa.arima.model as sm
from openpyxl import load_workbook

def main():
    pass # TODO: добавить отладчик и проверку работоспособности


def init_save_path():
    filepath = os.path.join(os.getcwd(), 'static', 'current_data', str(session['uid']))
    session['workfile'] = filepath


# Данные сохраняем в JSON. Так исключается проблема с движком openpyxl.
# TODO: нужна функция экспорта в формат excel после смены движка. Желательно новый (xlsx).
def save_data(data: pandas.DataFrame, postfix: str = ''):
    data.to_json(session['workfile'] + postfix + '.json')
    # data.to_excel(session['workfile'], engine='openpyxl')


def save_to_excel(data: pandas.DataFrame, filepath: str = 'unnamed.xlsx', index: bool = True, header: bool = True):
    data.to_excel(filepath, index=index, header=header)


def save_model(model: sm.ARIMAResults, filepath: str = ''):
    if filepath == '':
        filepath = os.path.join(os.getcwd(), 'static', 'models', str(session['uid']) + '_model.mdl')
    model.save(filepath)


def show_save_dialog(title: str = 'Сохранить файл', filetypes=(('Все файлы', '*.*'),)) -> str:
    import tkinter as tk
    from tkinter import filedialog
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    result = filedialog.asksaveasfilename(title=title, filetypes=filetypes)
    root.attributes('-topmost', False)
    root.destroy()
    return result


# ==== !!! Взято со stackoverflow.com
# TODO: заменить на свою реализацию
def append_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.

    @param filename: File path or existing ExcelWriter
                     (Example: '/path/to/file.xlsx')
    @param df: DataFrame to save to workbook
    @param sheet_name: Name of sheet which will contain DataFrame.
                       (default: 'Sheet1')
    @param startrow: upper left cell row to dump data frame.
                     Per default (startrow=None) calculate the last row
                     in the existing DF and write to the next row...
    @param truncate_sheet: truncate (remove and recreate) [sheet_name]
                           before writing DataFrame to Excel file
    @param to_excel_kwargs: arguments which will be passed to `DataFrame.to_excel()`
                            [can be a dictionary]
    @return: None

    Usage examples:

    >>> append_to_excel('d:/temp/test.xlsx', df)

    >>> append_to_excel('d:/temp/test.xlsx', df, header=None, index=False)

    >>> append_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False)

    >>> append_to_excel('d:/temp/test.xlsx', df, sheet_name='Sheet2',
                           index=False, startrow=25)

    (c) [MaxU](https://stackoverflow.com/users/5741205/maxu?tab=profile)
    """
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pandas.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()
